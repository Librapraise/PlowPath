import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_pricing_sheet():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # TAB 1: Cost Simulator
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Cost Simulator"
    ws1.views.sheetView[0].showGridLines = True
    
    # Colors
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    italic_font = Font(name="Calibri", size=9, italic=True)
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_bottom = Side(border_style="medium", color="1F4E78")
    double_bottom = Side(border_style="double", color="1F4E78")
    
    grid_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    header_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)
    total_border = Border(top=thin_line, bottom=double_bottom)
    
    # Title Block
    ws1["A1"] = "PlowPath Cost Simulator"
    ws1["A1"].font = title_font
    ws1.row_dimensions[1].height = 25
    
    # Section: Input Parameters
    ws1["A3"] = "Interactive Input Parameters"
    ws1["A3"].font = section_font
    
    inputs = [
        ("Number of Active Clients", 500, "Count of active customer properties (driveways/lots)"),
        ("Number of Active Drivers", 10, "Count of snowplow drivers running the mobile app"),
        ("Number of Storms per Month", 4, "Average active winter weather events in a month"),
        ("SMS Notification Ratio", 0.90, "Percentage of clients preferring SMS alerts (e.g. 0.90 for 90%)"),
        ("Voice Call Ratio", 0.10, "Percentage of clients preferring Voice Call alerts (e.g. 0.10 for 10%)"),
    ]
    
    for idx, (label, val, desc) in enumerate(inputs, start=4):
        ws1[f"A{idx}"] = label
        ws1[f"B{idx}"] = val
        ws1[f"C{idx}"] = desc
        ws1[f"A{idx}"].font = bold_font
        ws1[f"B{idx}"].font = bold_font
        ws1[f"B{idx}"].fill = input_fill
        ws1[f"B{idx}"].alignment = Alignment(horizontal="right")
        ws1[f"C{idx}"].font = italic_font
        ws1[f"A{idx}"].border = grid_border
        ws1[f"B{idx}"].border = grid_border
        ws1[f"C{idx}"].border = grid_border
        
        # Formats
        if "Ratio" in label:
            ws1[f"B{idx}"].number_format = "0%"
        else:
            ws1[f"B{idx}"].number_format = "#,##0"
            
    # Staging vs Production Selector Info
    ws1["A9"] = "Hosting & DB Tier Mode"
    ws1["B9"] = "Production"
    ws1["C9"] = "Enter 'Production' or 'Staging' to swap hosting cost baselines"
    ws1["A9"].font = bold_font
    ws1["B9"].font = bold_font
    ws1["B9"].fill = input_fill
    ws1["B9"].alignment = Alignment(horizontal="right")
    ws1["C9"].font = italic_font
    ws1["A9"].border = grid_border
    ws1["B9"].border = grid_border
    ws1["C9"].border = grid_border

    # Section: Calculated Monthly Costs
    ws1["A12"] = "Calculated Monthly Expenses"
    ws1["A12"].font = section_font
    
    headers = ["Cost Component", "Monthly Total", "Cost per Client", "Calculation Rule / Formula Detail"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws1.cell(row=13, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [2,3] else "left", vertical="center")
        cell.border = header_border
    ws1.row_dimensions[13].height = 24
    
    # Calculation rows
    # A14: Fixed/Idle Costs
    ws1["A14"] = "Platform Fixed Idle Cost"
    ws1["B14"] = '=IF(B9="Production", 41.14, 8.25)'
    ws1["C14"] = "=B14/B4"
    ws1["D14"] = "Baseline hosting, DB and developer account costs. (Staging = $8.25, Prod = $41.14)"
    
    # A15: Twilio SMS Costs
    ws1["A15"] = "Twilio SMS Notifications"
    ws1["B15"] = "=B4 * B7 * (B6 * 3) * 0.012"
    ws1["C15"] = "=B15/B4"
    ws1["D15"] = "Est. 3 texts per client per storm (pre-storm, en-route, done) at $0.012/msg (inc. carrier fees)"
    
    # A16: Twilio Voice Call Costs
    ws1["A16"] = "Twilio Voice Interactive Calls"
    ws1["B16"] = "=B4 * B8 * (B6 * 1.5) * 0.014"
    ws1["C16"] = "=B16/B4"
    ws1["D16"] = "Est. 1.5 minutes of interactive calling per storm per voice user at $0.014/min"
    
    # A17: Upstash Redis (Pay-as-you-go)
    ws1["A17"] = "Upstash Redis (Pay-As-You-Go)"
    ws1["B17"] = "=((B4 * 10 + B5 * 50) * B6 * 0.20) / 100000"
    ws1["C17"] = "=B17/B4"
    ws1["D17"] = "Serverless commands for tracking updates & Bull queues ($0.20 per 100k Redis commands)"
    
    # Formatting calc rows
    for r in range(14, 18):
        ws1[f"A{r}"].font = bold_font
        ws1[f"B{r}"].font = regular_font
        ws1[f"C{r}"].font = regular_font
        ws1[f"D{r}"].font = italic_font
        
        ws1[f"B{r}"].number_format = "$#,##0.00"
        ws1[f"C{r}"].number_format = "$#,##0.00"
        
        ws1[f"A{r}"].border = grid_border
        ws1[f"B{r}"].border = grid_border
        ws1[f"C{r}"].border = grid_border
        ws1[f"D{r}"].border = grid_border
        
    # Total row
    ws1["A18"] = "TOTAL ESTIMATED MONTHLY BILL"
    ws1["B18"] = "=SUM(B14:B17)"
    ws1["C18"] = "=SUM(C14:C17)"
    ws1["D18"] = "Combined platform idle + operational payload costs"
    
    ws1["A18"].font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    ws1["B18"].font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    ws1["C18"].font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    ws1["D18"].font = italic_font
    
    ws1["A18"].fill = green_fill
    ws1["B18"].fill = green_fill
    ws1["C18"].fill = green_fill
    
    ws1["B18"].number_format = "$#,##0.00"
    ws1["C18"].number_format = "$#,##0.00"
    
    ws1["A18"].border = total_border
    ws1["B18"].border = total_border
    ws1["C18"].border = total_border
    ws1["D18"].border = total_border

    # ----------------------------------------------------
    # TAB 2: Pricing Matrix
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Third-Party Pricing Matrix")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "PlowPath Detailed Pricing Matrix"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 25
    
    matrix_headers = ["Provider / Service", "Category", "Setup / One-time", "Idle Cost (Staging)", "Idle Cost (Production)", "Variable Cost Structure", "Notes / Limits"]
    for col_idx, text in enumerate(matrix_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    ws2.row_dimensions[3].height = 24
    
    matrix_data = [
        ("Apple Developer Program", "Mobile Distribution", 99.00, 8.25, 8.25, "None", "Annual Developer license, paid upfront ($99/year)"),
        ("Google Play Console", "Mobile Distribution", 25.00, 0.00, 0.00, "None", "One-time registration fee ($25)"),
        ("Transistor Background SDK", "Background GPS", 0.00, 0.00, 0.00, "None", "Optional production upgrade: $349 one-time. Default is free MIT library"),
        ("Fly.io (API)", "Cloud Hosting", 0.00, 0.00, 3.19, "None", "Free tier (256MB VM) in staging; 512MB RAM VM in production"),
        ("Self-Hosted OSRM (Fly.io)", "Map Routing Engine", 0.00, 0.00, 7.20, "None", "OSRM docker container on 1GB RAM + 10GB persistent storage"),
        ("Neon Postgres", "Database", 0.00, 0.00, 19.00, "None", "Free tier staging (0.5GB limit); paid Launch plan for production scale"),
        ("Upstash Redis", "Queue / Caching", 0.00, 0.00, 0.00, "$0.20 / 100k commands", "Pay-as-you-go plan. Zero idle costs. Staging tier free up to 10k/day"),
        ("Twilio SMS & Voice", "Communications", 59.00, 0.00, 3.50, "SMS: $0.012/msg, Voice: $0.014/min", "Prod: $1.15/mo number + $2.00/mo A2P campaign. Surcharges included in msg rate"),
        ("Mapbox Geocoding", "Address Search", 0.00, 0.00, 0.00, "$0.75 / 1,000 requests", "Includes 100,000 free temporary geocodes per month"),
        ("Resend Email", "Transaction Email", 0.00, 0.00, 0.00, "None", "Free tier covers 3,000 emails/month (100/day)"),
        ("Sentry", "Error Tracker", 0.00, 0.00, 0.00, "None", "Developer tier covers 5,000 alerts/month for free"),
        ("Cloudflare Pages", "Web Hosting", 0.00, 0.00, 0.00, "None", "Free static hosting for web dashboard (Vite dist)"),
    ]
    
    for r_idx, row_data in enumerate(matrix_data, start=4):
        ws2.row_dimensions[r_idx].height = 20
        # Determine Zebra stripe
        bg_stripe = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = grid_border
            if bg_stripe.fill_type:
                cell.fill = bg_stripe
                
            # Alignment & number formats
            if c_idx in [3, 4, 5]:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 1:
                cell.font = bold_font
            elif c_idx == 6:
                cell.alignment = Alignment(horizontal="center" if "None" in str(val) else "left")

    # ----------------------------------------------------
    # Formatting adjustments (Column Widths)
    # ----------------------------------------------------
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and ('$' in cell.number_format or '%' in cell.number_format):
                    # Pad length estimate for formatted values
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Specific manually-adjusted columns
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 75
    
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 22
    ws2.column_dimensions["F"].width = 28
    ws2.column_dimensions["G"].width = 65

    # Save to disk
    wb.save("c:/Users/folfe/Downloads/plowpath/docs/PlowPath_Pricing_Calculator.xlsx")
    print("Excel Spreadsheet created successfully at docs/PlowPath_Pricing_Calculator.xlsx")

if __name__ == "__main__":
    create_pricing_sheet()
